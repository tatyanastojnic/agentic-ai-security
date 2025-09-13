import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
from datetime import datetime

def generate_excel_report(results):
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"llm_security_report_{timestamp}.xlsx"
    wb = openpyxl.Workbook()
    
    # Summary sheet
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary.append(["LLM", "Total Tests", "Passed", "Failed", "Pass Rate (%)"])
    llms = set(r["LLM"] for r in results)
    for llm in llms:
        llm_results = [r for r in results if r["LLM"] == llm]
        total = len(llm_results)
        passed = sum(1 for r in llm_results if r["Status"] == "PASS")
        failed = total - passed
        ws_summary.append([llm, total, passed, failed, f"{passed/total*100:.1f}%"])
    
    # Sheets per LLM
    for llm in llms:
        ws = wb.create_sheet(title=llm)
        headers = ["ID", "Type", "Name", "Prompt", "Response", "Score", "Expected_Behavior", "Status"]
        ws.append(headers)
        for row_data in [r for r in results if r["LLM"] == llm]:
            ws.append([
                row_data["ID"], row_data["Type"], row_data["Name"], row_data["Prompt"],
                row_data["Response"], row_data["Score"], row_data["Expected_Behavior"], row_data["Status"]
            ])
            last_row = ws.max_row
            status_cell = ws[f"H{last_row}"]
            if row_data["Status"] == "PASS":
                status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                status_cell.font = Font(bold=True)
            else:
                status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                status_cell.font = Font(bold=True)
        for col in ws.columns:
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col) + 2
            ws.column_dimensions[get_column_letter(col[0].column)].width = max_length
    
    wb.save(filename)
    print(f"Multi-LLM Excel report saved: {filename}")
